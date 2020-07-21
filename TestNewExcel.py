
from openpyxl import load_workbook

wb = load_workbook('北京9月考勤数据.xls')
print(wb.sheetnames)